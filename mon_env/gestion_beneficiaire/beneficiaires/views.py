
from .models import Beneficiaire, Conjoint, Enfant, OrigineBeneficiaire, InformationOperation
from .forms import BeneficiaireForm, ConjointForm, EnfantForm, OrigineBeneficiaireForm, InformationOperationForm
import openpyxl
from django.core.paginator import Paginator
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from django.contrib.auth import login, authenticate
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.models import User
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.contrib.auth import logout 
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash

@login_required
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, 'Your password was successfully updated!')
            return redirect('change_password')
        else:
            messages.error(request, 'Please correct the error below.')
    else:
        form = PasswordChangeForm(user=request.user)
    return render(request, 'pages/change_password.html', {'form': form})

def logout_user(request):
    logout(request)
    messages.success(request, "Logged out Successfully!")
    return redirect('login')

def login_view(request):
    error_message = ''
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                # Redirect to a success page or the gestion beneficiaire
                return redirect('gestion_bn')  # Replace 'gestion_bn' with your desired URL name
            else:
                error_message = "Invalid username or password."
        else:
            error_message = "Invalid username or password."
    else:
        form = AuthenticationForm()

    return render(request, 'pages/login.html', {'form': form, 'error_message': error_message})
    # If request method is GET (initial page load)

def signup_view(request):
    
    if request.method == 'POST':
        username = request.POST['username']
        email = request.POST['email']
        password1 = request.POST['password1']
        password2 = request.POST['password2']

        myuser = User.objects.create_user(username,email,password1)
        myuser.save()

        messages.success(request, "Your Account has been successfully created.")
        return redirect('login')
    else:
        form = UserCreationForm()

    return render(request, 'pages/signup.html', {'form': form})


def export_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="search_results.xlsx"'

    beneficiaires = Beneficiaire.objects.all()

    # Applying filters based on the GET parameters
    cin = request.GET.get('cin')
    nom = request.GET.get('nom')
    prenom = request.GET.get('prenom')
    num_pv = request.GET.get('num_pv')
    nom_operation = request.GET.get('nom_operation')
    cin_conjoint = request.GET.get('cin_conjoint')
    nom_conjoint = request.GET.get('nom_conjoint')
    prenom_conjoint = request.GET.get('prenom_conjoint')
    nom_enfant = request.GET.get('nom_enfant')
    prenom_enfant = request.GET.get('prenom_enfant')
    num_lot = request.GET.get('num_lot')

    if cin:
        beneficiaires = beneficiaires.filter(cin__icontains=cin)
    if nom:
        beneficiaires = beneficiaires.filter(nom__icontains=nom)
    if prenom:
        beneficiaires = beneficiaires.filter(prenom__icontains=prenom)
    if num_pv:
        beneficiaires = beneficiaires.filter(informationoperation__num_pv__icontains=num_pv)
    if nom_operation:
        beneficiaires = beneficiaires.filter(informationoperation__nom_operation__icontains=nom_operation)
    if num_lot:
        beneficiaires = beneficiaires.filter(informationoperation__num_lot__icontains=num_lot)

# Filtering conjoint and enfant separately
    if cin_conjoint:
        conjoint = Conjoint.objects.filter(cin__icontains=cin_conjoint)
        beneficiaires = beneficiaires.filter(conjoint__in=conjoint)
    if nom_conjoint:
        conjoint = Conjoint.objects.filter(nom__icontains=nom_conjoint)
        beneficiaires = beneficiaires.filter(conjoint__in=conjoint)
    if prenom_conjoint:
        conjoint = Conjoint.objects.filter(prenom__icontains=prenom_conjoint)
        beneficiaires = beneficiaires.filter(conjoint__in=conjoint)

    if nom_enfant:
        enfant = Enfant.objects.filter(nom__icontains=nom_enfant)
        beneficiaires = beneficiaires.filter(enfant__in=enfant)
    if prenom_enfant:
        enfant = Enfant.objects.filter(prenom__icontains=prenom_enfant)
        beneficiaires = beneficiaires.filter(enfant__in=enfant) 
    # Create Excel workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write headers
    headers = ['CIN du bénéficiaire', 'Nom du bénéficiaire', 'Prénom du bénéficiaire',
               'Nom de l\'opération', 'Opérateur', 'Date d\'affectation', 'N° de Ressencement']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # Write data rows
    for row_num, beneficiaire in enumerate(beneficiaires, 2):
        ws.cell(row=row_num, column=1, value=beneficiaire.cin)
        ws.cell(row=row_num, column=2, value=beneficiaire.nom)
        ws.cell(row=row_num, column=3, value=beneficiaire.prenom)
        info_operations = beneficiaire.informationoperation_set.all()
        if info_operations.exists():
            ws.cell(row=row_num, column=4, value=info_operations[0].nom_operation)
            ws.cell(row=row_num, column=5, value=info_operations[0].operateur)
            ws.cell(row=row_num, column=6, value=info_operations[0].date_affectation)
            ws.cell(row=row_num, column=7, value=info_operations[0].num_ressencement)

    # Save the Excel file
    wb.save(response)
    return response

# Create your views here.
def gestion_bn(request):
    beneficiaires = Beneficiaire.objects.all()
    info_operations = InformationOperation.objects.all()
    page = Paginator(beneficiaires, 10)
    page_list = request.GET.get('page')
    page = page.get_page(page_list)
    context = {
        'page': page,
        'info_operations': info_operations,
    }

    return render(request, 'pages/gestion_bn.html', context)


def add_beneficiary(request):
    error_message = None
    cin = request.POST.get('cin')
    if request.method == 'POST':
        if Beneficiaire.objects.filter(cin=cin).exists():
            error_message = "Un bénéficiaire avec le même 'cin' existe déjà."
        else:
            nom = request.POST.get('nom')
            prenom = request.POST.get('prenom')
            spouse_cin = request.POST.get('spouse_cin')
            spouse_nom = request.POST.get('spouse_nom')
            spouse_prenom = request.POST.get('spouse_prenom')
            child_nom = request.POST.get('child_nom')
            child_prenom = request.POST.get('child_prenom')
            ville = request.POST.get('ville')
            commune = request.POST.get('commune')
            douar = request.POST.get('douar')
            num_pv = request.POST.get('num_pv')
            operateur = request.POST.get('operateur')
            nom_operation = request.POST.get('nom_operation')
            num_ressencement = request.POST.get('num_ressencement')
            date_ressencement = request.POST.get('date_ressencement')
            date_affectation = request.POST.get('date_affectation')
            type_intervention = request.POST.get('type_intervention')
            prix_produit = request.POST.get('prix_produit')
            subvention_fshiu = request.POST.get('subvention_fshiu')
            date_demolition = request.POST.get('date_demolition')
            num_lot = request.POST.get('num_lot')
            fichier_joint = request.FILES.get('fichier_joint')
            observation = request.POST.get('observation')
            # Create Beneficiaire instance
            beneficiary = Beneficiaire(cin=cin, nom=nom, prenom=prenom)
            beneficiary.save()

            # Create Conjoint instance if spouse information provided
            if spouse_cin and spouse_nom and spouse_prenom:
                spouse = Conjoint(cin=spouse_cin, nom=spouse_nom, prenom=spouse_prenom, beneficiaire=beneficiary)
                spouse.save()

            # Create Enfant instance if child information provided
            if child_nom and child_prenom:
                child = Enfant(nom=child_nom, prenom=child_prenom, beneficiaire=beneficiary)
                child.save()

            # Create OrigineBeneficiaire instance
            origine = OrigineBeneficiaire(ville=ville, commune=commune, douar=douar, beneficiaire=beneficiary)
            origine.save()

            # Create InformationOperation instance
            info_operation = InformationOperation(
                num_pv=num_pv,
                operateur=operateur,
                nom_operation=nom_operation,
                num_ressencement=num_ressencement,
                date_ressencement=date_ressencement,
                date_affectation=date_affectation,
                type_intervention=type_intervention,
                prix_produit=prix_produit,
                subvention_fshiu=subvention_fshiu,
                date_demolition=date_demolition,
                num_lot=num_lot,
                fichier_joint=fichier_joint,
                observation=observation,
                beneficiaire=beneficiary
            )
            info_operation.save()

            return redirect('gestion_bn')

    return render(request, 'pages/form_bn.html',{'error_message': error_message})

def update_beneficiary(request,cin):
    try:
        # Fetch the existing beneficiary instance from the database
        beneficiary = get_object_or_404(Beneficiaire, cin=cin)
        spouse = Conjoint.objects.filter(beneficiaire=beneficiary).first()
        child = Enfant.objects.filter(beneficiaire=beneficiary).first()
        origine = OrigineBeneficiaire.objects.filter(beneficiaire=beneficiary).first()
        info_operation = InformationOperation.objects.filter(beneficiaire=beneficiary).first()

    except Beneficiaire.DoesNotExist:
        # Handle the case when the beneficiary with the given ID doesn't exist
        # You can redirect to an error page or return an error response here
        pass

    if request.method == 'POST':
        cin = request.POST.get('cin')
        nom = request.POST.get('nom')
        prenom = request.POST.get('prenom')
        spouse_cin = request.POST.get('spouse_cin')
        spouse_nom = request.POST.get('spouse_nom')
        spouse_prenom = request.POST.get('spouse_prenom')
        child_nom = request.POST.get('child_nom')
        child_prenom = request.POST.get('child_prenom')
        ville = request.POST.get('ville')
        commune = request.POST.get('commune')
        douar = request.POST.get('douar')
        num_pv = request.POST.get('num_pv')
        operateur = request.POST.get('operateur')
        nom_operation = request.POST.get('nom_operation')
        num_ressencement = request.POST.get('num_ressencement')
        date_ressencement = request.POST.get('date_ressencement')
        date_affectation = request.POST.get('date_affectation')
        
        type_intervention = request.POST.get('type_intervention')
        prix_produit = request.POST.get('prix_produit')
        subvention_fshiu = request.POST.get('subvention_fshiu')
        date_demolition = request.POST.get('date_demolition')
        num_lot = request.POST.get('num_lot')
        fichier_joint = request.FILES.get('fichier_joint')
        observation = request.POST.get('observation')

       # Update the attributes of the existing beneficiary instance
        beneficiary.cin = cin
        beneficiary.nom = nom
        beneficiary.prenom = prenom
        beneficiary.save()

        # Update spouse information if provided
        if spouse_cin and spouse_nom and spouse_prenom:
            spouse, _ = Conjoint.objects.get_or_create(beneficiaire=beneficiary)
            spouse.cin = spouse_cin
            spouse.nom = spouse_nom
            spouse.prenom = spouse_prenom
            spouse.save()

        # Update child information if provided
        if child_nom and child_prenom:
            child, _ = Enfant.objects.get_or_create(beneficiaire=beneficiary)
            child.nom = child_nom
            child.prenom = child_prenom
            child.save()

        # Update OrigineBeneficiaire information
        origine, _ = OrigineBeneficiaire.objects.get_or_create(beneficiaire=beneficiary)
        origine.ville = ville
        origine.commune = commune
        origine.douar = douar
        origine.save()

        # Update InformationOperation information
        info_operation, _ = InformationOperation.objects.get_or_create(beneficiaire=beneficiary)
        info_operation.num_pv = num_pv
        info_operation.operateur = operateur
        info_operation.nom_operation = nom_operation
        info_operation.num_ressencement = num_ressencement
        info_operation.date_ressencement = date_ressencement
        info_operation.date_affectation = date_affectation
        info_operation.type_intervention = type_intervention
        info_operation.prix_produit = prix_produit
        info_operation.subvention_fshiu = subvention_fshiu
        info_operation.date_demolition = date_demolition
        info_operation.num_lot = num_lot
        info_operation.fichier_joint = fichier_joint
        info_operation.observation = observation
        info_operation.save()

        return redirect('gestion_bn')

    return render(request, 'pages/modify_bn.html', {
        'beneficiary': beneficiary,
        'spouse': spouse,
        'child': child,
        'origine': origine,
        'info_operation': info_operation,
        })

def delete_beneficiary(request, cin):
    beneficiary = get_object_or_404(Beneficiaire, cin=cin)

    if request.method == 'POST':
        beneficiary.delete()
        return redirect('gestion_bn')

    return render(request, 'pages/delete_bn.html', {'beneficiary': beneficiary})

def detail_beneficiary(request, cin):
    beneficiary = get_object_or_404(Beneficiaire, cin=cin)
    spouse = Conjoint.objects.filter(beneficiaire=beneficiary).first()
    children = Enfant.objects.filter(beneficiaire=beneficiary)
    origine_beneficiaire = OrigineBeneficiaire.objects.get(beneficiaire=beneficiary)
    information_operation = InformationOperation.objects.get(beneficiaire=beneficiary)

    return render(request, 'pages/detail_bn.html', {
        'beneficiary': beneficiary,
        'spouse': spouse,
        'children': children,
        'origine_beneficiaire': origine_beneficiaire,
        'information_operation': information_operation,
    })

def base(request):
    return render(request, 'base.html')

def search_beneficiaires(request):
    if request.method == 'GET':
        cin = request.GET.get('cin')
        nom = request.GET.get('nom')
        prenom = request.GET.get('prenom')
        num_pv = request.GET.get('num_pv')
        nom_operation  = request.GET.get('nom_operation')
        cin_conjoint = request.GET.get('cin_conjoint')
        nom_conjoint = request.GET.get('nom_conjoint')
        prenom_conjoint = request.GET.get('prenom_conjoint')
        nom_enfant = request.GET.get('nom_enfant')
        prenom_enfant = request.GET.get('prenom_enfant')
        num_lot = request.GET.get('num_lot')

        beneficiaires = Beneficiaire.objects.all()

        if cin:
            beneficiaires = beneficiaires.filter(cin__icontains=cin)
        if nom:
            beneficiaires = beneficiaires.filter(nom__icontains=nom)
        if prenom:
            beneficiaires = beneficiaires.filter(prenom__icontains=prenom)
        if num_pv:
            beneficiaires = beneficiaires.filter(informationoperation__num_pv__icontains=num_pv)
        if nom_operation :
            beneficiaires = beneficiaires.filter(informationoperation__nom_operation__icontains=nom_operation )
        if num_lot:
            beneficiaires = beneficiaires.filter(informationoperation__num_lot__icontains=num_lot)


        # Filtering conjoint and enfant separately
        if cin_conjoint:
            conjoint = Conjoint.objects.filter(cin__icontains=cin_conjoint)
            beneficiaires = beneficiaires.filter(conjoint__in=conjoint)
        if nom_conjoint:
            conjoint = Conjoint.objects.filter(nom__icontains=nom_conjoint)
            beneficiaires = beneficiaires.filter(conjoint__in=conjoint)
        if prenom_conjoint:
            conjoint = Conjoint.objects.filter(prenom__icontains=prenom_conjoint)
            beneficiaires = beneficiaires.filter(conjoint__in=conjoint)

        if nom_enfant:
            enfant = Enfant.objects.filter(nom__icontains=nom_enfant)
            beneficiaires = beneficiaires.filter(enfant__in=enfant)
        if prenom_enfant:
            enfant = Enfant.objects.filter(prenom__icontains=prenom_enfant)
            beneficiaires = beneficiaires.filter(enfant__in=enfant)
        

        return render(request, 'pages/search.html', {'beneficiaires': beneficiaires})
    else:
        return render(request, 'pages/search.html')
